import csv
import openpyxl
import numpy as np
import datetime

wb = openpyxl.load_workbook("TTLC Tracker.xlsx")
ws = wb["Input"]

def internal(input):
    internal_points = []
    internal_points.append(internal_attendance(input[0]))
    internal_points.append(internal_MOTM(input[1]))
    internal_points.append(internal_setup(input[2]))
    return internal_points

def internal_attendance(value):
    return value * 0.25

def internal_MOTM(value):
    return value * 3

def internal_setup(value):
    return value * (0.9 ** (value - 1))

def engagement(input):
    engagement_points = []
    engagement_points.append(engagement_chalking(input[0]))
    engagement_points.append(engagement_recruit(input[1]))
    engagement_points.append(engagement_social(input[2]))
    return engagement_points

def engagement_chalking(value):
    return value * 0.5

def engagement_recruit(value):
    return value * 0.5

def engagement_social(value):
    return value * (0.67 ** (value - 1))

def communications(input):
    comm_points = []
    comm_points.append(communications_social(input[0]))
    comm_points.append(communications_MDStories(input[1]))
    comm_points.append(communications_PenPal(input[2]))
    return comm_points

def communications_social(value):
    return value * 0.5

def communications_MDStories(value):
    return value * 2

def communications_PenPal(value):
    return value * 1.5

def fiscal(input):
    fiscal_points = []
    fiscal_points.append(fiscal_social(input[0]))
    fiscal_points.append(fiscal_push(input[1]))
    fiscal_points.append(fiscal_canning(input[2]))
    return fiscal_points

def fiscal_social(value):
    return value * (0.67 ** (value - 1))

def fiscal_push(value):
    return value * 0.5

def fiscal_canning(value):
    return value * (0.5)

def run():
    today = datetime.date.today()
    with open(str(today) + "-Report.csv", "w+") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(['Name', 'Committee', 'Points', 'Threshold'])
        for i in range(3, 250):
            try:
                row = []
                if (ws['A' + str(i)].value == None):
                    break
                row.append(ws['A' + str(i)].value)
                row.append(ws['B' + str(i)].value)

                i_input = [ws[('D' + str(i))].value, ws[('E' + str(i))].value, ws[('F' + str(i))].value]
                e_input = [ws[('G' + str(i))].value, ws[('H' + str(i))].value, ws[('I' + str(i))].value]
                c_input = [ws[('J' + str(i))].value, ws[('J' + str(i))].value, ws[('L' + str(i))].value]
                f_input = [ws[('M' + str(i))].value, ws[('N' + str(i))].value, ws[('O' + str(i))].value]
                sum = 0
                sum += np.sum(internal(i_input))
                sum += np.sum(engagement(e_input))
                sum += np.sum(communications(c_input))
                sum += np.sum(fiscal(f_input))
                row.append(sum)
                if (sum > 17.00):
                    row.append('Yes')
                else:
                    row.append('No')
                writer.writerow(row)
            except Exception as e:
                print(e)
                break


    print("Report Generated Successfully")

run()